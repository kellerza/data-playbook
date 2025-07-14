"""Read helpers."""

from __future__ import annotations
import logging
from collections.abc import Generator, Sequence
from json import dumps
from pathlib import Path
from typing import Any

import attrs
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.worksheet.worksheet import Worksheet

from dataplaybook import PathStr, RowData, RowDataGen, Tables, task

_LOGGER = logging.getLogger(__name__)


@attrs.define
class Column:
    """An Excel column definition."""

    name: str
    source: int | str = ""
    """From header/col index."""
    width: float = 9

    @classmethod
    def from_old(
        cls, old: list[dict] | dict[str, dict[str, str]] | Any
    ) -> list[Column] | None:
        """Convert old columns to new."""
        if isinstance(old, list):
            # old write format
            return [Column(name=c["name"], width=c.get("width", 9)) for c in old]
        if isinstance(old, dict):
            # old read format
            return [
                Column(name=k, source=v["col"] if "col" in v else v.get("from") or "")
                for k, v in old.items()
            ]
        if old:
            raise ValueError(f"Bad column definition: {old}")
        return None


@attrs.define
class Sheet:
    """An Excel sheet definition."""

    name: str = ""
    source: str = ""
    """Source sheet. May be '*' for the default/active sheet"""
    header: int = 0
    columns: list[Column] | None = attrs.field(default=None)

    @classmethod
    def from_old(cls, *headers: dict[str, Any]) -> list[Sheet]:
        """Convert old headers to sheets."""
        res = []
        for sheet in headers:
            if isinstance(sheet, dict) and "sheet" in sheet:
                res.append(  # write
                    Sheet(
                        name=sheet["sheet"],
                        columns=Column.from_old(sheet.get("columns")),
                    )
                )
                continue
            if isinstance(sheet, dict) and "target" in sheet and "name" in sheet:
                res.append(  # read def
                    Sheet(
                        name=sheet["target"],
                        source=sheet["name"],
                        header=sheet.get("header") or 0,
                        columns=Column.from_old(sheet.get("columns")),
                    )
                )
                continue
            raise ValueError(f"Bad sheet definition: {sheet}")
        return res


@task
def read_excel(
    *, tables: Tables, file: PathStr, sheets: list[Sheet] | None = None
) -> list[str]:
    """Read excel file using openpyxl.

    If no sheets are specified, all sheets are read and names returned.
    """
    wbk = openpyxl.load_workbook(file, read_only=True, data_only=True)
    _LOGGER.debug("Loaded workbook %s.", file)

    if sheets is None:
        sheets = [Sheet(name=n) for n in wbk.sheetnames]
    if not sheets:
        sheets = []

    res: list[str] = []

    for sht in sheets:
        name = sht.source or sht.name
        # default_sheet = *
        the_sheet = wbk.active if name == "*" else wbk[name]
        if not the_sheet:
            _LOGGER.error("Sheet %s not found", name)
            tables[name] = []
            continue
        tbl = _sheet_read(the_sheet, sht)
        res.append(sht.name)
        tables[sht.name] = tbl

    return res


def _sheet_read(_sheet: Worksheet, shdef: Sheet) -> list[RowData]:
    """Read a sheet and return a table."""
    res = list(_sheet_yield_rows(_sheet, shdef))
    _LOGGER.debug("Read %s rows from sheet %s", len(res), _sheet.title)
    return res


def _column_map(
    columns: list[Column] | None, header_row: Sequence[str]
) -> Generator[tuple[int, str, Column | None], None, None]:
    """List of (idx, nme, val)."""
    if not columns:
        for idx, key in enumerate(header_row):
            if key:
                yield (idx, key, None)
        return

    for col in columns:
        if isinstance(col.source, int) and col.source >= 0:
            yield (col.source, col.name, col)
            continue
        idx = list(header_row).index(str(col.source))
        if col.source and idx < 0:
            _LOGGER.error("%s not found in header %s", col.source, list(header_row))
            raise ValueError(f"{col.source} not found in header")
        if idx < 0:
            raise ValueError(f"Bad column definition: {col.name}")
        yield (idx, col.name, col)


def _sheet_yield_rows(_sheet: Worksheet, shdef: Sheet) -> RowDataGen:
    """Read the sheet and yield the rows."""
    rows = _sheet.rows
    header = shdef.header
    try:
        while header > 0:
            next(rows)
            header -= 1
        header_row = [str(cell.value) if cell.value else "" for cell in next(rows)]
    except StopIteration:
        header_row = []
    while len(header_row) > 1 and header_row[-1] is None:
        header_row.pop()
    _LOGGER.debug("Header row: %s", header_row)

    colmap = list(_column_map(shdef.columns, header_row))

    row = None
    nocnt = 0
    while True:
        try:
            row = next(rows)
        except StopIteration:
            return
        record = {}
        for idx, key, _ in colmap:
            _cv = row[idx].value
            record[key] = _cv.strip() if isinstance(_cv, str) else _cv
        if record:
            yield record
            nocnt = 0
        else:
            nocnt += 1
            if nocnt > 1000:
                return


def _get_filename(filename: PathStr) -> str:
    """Get a filename to write to."""
    path = Path(filename)
    try:
        if path.exists() and path.is_file():
            path.unlink()
        return str(filename)
    except OSError:  # Open in Excel?
        for idx in range(50):
            newf = path.with_stem(path.stem + f" {idx}")
            if not newf.is_file():
                _LOGGER.warning("File %s locked, saving as %s", filename, newf)
                return str(newf)
        raise


def _fmt(obj: Any) -> str:
    """Format an object for Excel."""
    if callable(obj):
        return str(obj)
    try:
        if isinstance(obj, list | dict | tuple):
            return dumps(obj)
    except TypeError:
        return str(obj)

    # openpyxl's _bind_value in cell.py doesn't use isinstance
    if isinstance(obj, str) and type(obj) != str:  # noqa: E721
        return str(obj)

    return obj


@task
def write_excel(  # noqa: PLR0912
    *,
    tables: Tables,
    file: PathStr,
    include: list[str] | None = None,
    sheets: list[Sheet] | None = None,
    ensure_string: bool = False,
) -> None:
    """Write an excel file."""
    if sheets:
        if not all(isinstance(s, Sheet) for s in sheets):
            raise ValueError("sheets must be a list of Sheet objects")

    sheets = sheets or []
    wbk = openpyxl.Workbook()

    if not include:
        include = list(tables.keys())

    # prep headers
    sheet_lookup = {i.name: i.columns for i in sheets}

    # Remove default sheet
    wbk.remove(wbk["Sheet"])

    for table_name in include:
        # Skip empty tables
        if not tables[table_name]:
            continue

        wsh: Worksheet = wbk.create_sheet(table_name)

        if table_name not in tables:
            _LOGGER.warning("Could not save table %s", table_name)
            continue

        hdr: dict[str, int] = {}

        # Get initial header from headers input
        if cols := sheet_lookup.get(table_name):
            for col in cols:
                hdr[col.name] = 1
                wsh.column_dimensions[get_column_letter(len(hdr))].width = (
                    col.width or 9
                )

        # Ensure we get then all
        for row in tables[table_name]:
            for _hdr in row.keys():
                hdr[str(_hdr)] = 1
        hdrk = list(hdr.keys())
        wsh.append(hdrk)

        debugs = 4
        for row in tables[table_name]:
            erow = [_fmt(row.get(h)) for h in hdrk]
            try:
                wsh.append(erow)
            except IllegalCharacterError:
                newrow = [
                    ILLEGAL_CHARACTERS_RE.sub(" ", v) if isinstance(v, str) else v
                    for v in erow
                ]
                wsh.append(newrow)
            except ValueError as exc:
                debugs -= 1
                if debugs > 0:
                    _LOGGER.warning(
                        "Error writing %s, hdrs: %s - %s", list(erow), hdrk, exc
                    )
                wsh.append([str(row.get(h, "")) for h in hdrk])
        if debugs < 0:
            _LOGGER.warning("Total %s errors", 2 - debugs)

    wbk.save(_get_filename(file))  # Write to disk

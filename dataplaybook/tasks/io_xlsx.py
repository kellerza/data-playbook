"""Read helpers."""

import logging
import os
from json import dumps
from typing import Any, Sequence

import openpyxl
import voluptuous as vol
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import dataplaybook.config_validation as cv
from dataplaybook import RowData, RowDataGen, Tables, task_validate
from dataplaybook.utils import ensure_list as _ensure_list

_LOGGER = logging.getLogger(__name__)


@task_validate(
    validator=vol.Schema(
        schema={
            vol.Required("tables"): cv.ensure_tables,
            vol.Required("file"): str,
            vol.Optional("sheets"): [
                vol.Schema(
                    {
                        vol.Optional("name", default=None): vol.Any(str, None),
                        vol.Optional("header"): int,
                        vol.Optional("columns", default=None): vol.Any(
                            None,
                            vol.Schema(
                                {
                                    str: vol.Schema(
                                        {
                                            vol.Optional("from"): str,
                                            vol.Optional("col"): int,
                                        }
                                    )
                                }
                            ),
                        ),
                        vol.Required("target"): str,
                    }
                )
            ],
        }
    )
)
def read_excel(
    *, tables: Tables, file: str, sheets: list[dict[str, Any]] | None = None
) -> list[str]:
    """Read excel file using openpyxl.

    If no sheets are specified, all sheets are read and names returned."""
    wbk = openpyxl.load_workbook(file, read_only=True, data_only=True)
    _LOGGER.debug("Loaded workbook %s.", file)

    if sheets is None:
        sheets = [{"name": n, "target": n} for n in wbk.sheetnames]
    if not sheets:
        sheets = []

    for sht in sheets:
        name = sht.get("name") or sht["target"]
        # default_sheet = *
        the_sheet = wbk.active if name == "*" else wbk[name]
        tbl = _sheet_read(  # pylint: disable=protected-access
            the_sheet, sht.get("columns"), sht.get("header", 0)
        )
        tables[sht["target"]] = tbl

    return [s["target"] for s in sheets]


def _sheet_read(
    _sheet: Worksheet, columns: dict | None = None, header: int = 0
) -> list[RowData]:
    """Read a sheet and return a table."""
    res = list(_sheet_yield_rows(_sheet, columns, header))
    _LOGGER.debug("Read %s rows from sheet %s", len(res), _sheet.title)
    return res


def _column_map(
    columns: dict | None, header_row: Sequence[str]
) -> Sequence[tuple[int, str, Any | None]]:
    """List of (idx, nme, val)."""
    res = []
    if not columns:
        for idx, key in enumerate(header_row):
            if key:
                res.append((idx, str(key), None))
        return res

    for nme, val in columns.items():
        if "from" in val:
            fromc = str(val["from"])
            if fromc not in header_row:
                _LOGGER.error("%s not found in header %s", fromc, list(header_row))
                raise ValueError(f"{fromc} not found in header")
            res.append((list(header_row).index(fromc), nme, val))
        elif "col" in val:
            res.append((val["col"], nme, val))
        else:
            raise ValueError(f"Bad column definition: {val}")
    return res


def _sheet_yield_rows(
    _sheet: Worksheet, columns: dict | None = None, header: int = 0
) -> RowDataGen:
    """Read the sheet and yield the rows."""
    rows = _sheet.rows
    try:
        while header > 0:
            next(rows)
            header -= 1
        header_row = [cell.value for cell in next(rows)]
    except StopIteration:
        header_row = []
    while len(header_row) > 1 and header_row[-1] is None:
        header_row.pop()
    _LOGGER.debug("Header row: %s", header_row)

    colmap = _column_map(columns, header_row)

    row = None
    nocnt = 0
    while True:
        try:
            row = next(rows)
        except StopIteration:
            return
        record = {}
        for idx, key, _ in colmap:
            _cv = row[idx].value
            record[key] = _cv.strip() if isinstance(_cv, str) else _cv
        if record:
            yield record
            nocnt = 0
        else:
            nocnt += 1
            if nocnt > 1000:
                return


def _get_filename(filename: str) -> str:
    """Get a filename to write to."""
    try:
        if os.path.isfile(filename):
            os.remove(filename)
        return filename
    except OSError:  # Open in Excel?
        _parts = list(os.path.splitext(filename))
        _parts[0] += " "
        for idx in range(50):
            newname = str(idx).join(_parts)
            if not os.path.isfile(newname):
                _LOGGER.warning("File %s locked, saving as %s", filename, newname)
                return newname
        raise


def _fmt(obj: Any) -> str:
    """Format an object for Excel."""
    if callable(obj):
        return str(obj)
    try:
        if isinstance(obj, (list, dict, tuple)):
            return dumps(obj)
    except TypeError:
        return str(obj)

    # openpyxl's _bind_value in cell.py doesn't use isinstance

    if (
        isinstance(obj, str) and type(obj) != str  # noqa: E721 # pylint: disable=unidiomatic-typecheck
    ):
        return str(obj)

    return obj


@task_validate(
    validator=vol.Schema(
        {
            vol.Required("tables"): cv.ensure_tables,
            vol.Required("file"): cv.endswith(".xlsx"),
            vol.Optional("include"): vol.All(_ensure_list, [str]),
            vol.Optional("header", default=[]): vol.All(_ensure_list, [str]),
            vol.Optional("headers", default=[]): vol.All(_ensure_list, [object]),
        }
    )
)
def write_excel(
    *,
    tables: Tables,
    file: str,
    include: list[str] | None = None,
    header: list[str] | None = None,
    headers: list[Any] | None = None,
    ensure_string: bool = False,
) -> None:
    """Write an excel file."""
    header = header or []
    headers = headers or []
    wbk = openpyxl.Workbook()

    if not include:
        include = list(tables.keys())

    # prep headers
    headers_lookup = {i["sheet"]: i["columns"] for i in headers}

    # Remove default sheet
    wbk.remove(wbk["Sheet"])

    for table_name in include:
        # Skip empty tables
        if not tables[table_name]:
            continue

        wsh = wbk.create_sheet(table_name)

        if table_name not in tables:
            _LOGGER.warning("Could not save table %s", table_name)
            continue

        hdr: dict[str, int] = {}

        # Old style headers (applies to all sheets)
        for _hdr in header:
            hdr[_hdr] = 1

        # Get initial header from headers input
        if table_name in headers_lookup:
            for col in headers_lookup[table_name]:
                hdr[col["name"]] = 1
                wsh.column_dimensions[get_column_letter(len(hdr))].width = col.get(
                    "width", 9
                )

        # Ensure we get then all
        for row in tables[table_name]:
            for _hdr in row.keys():
                hdr[str(_hdr)] = 1
        hdrk = list(hdr.keys())
        wsh.append(hdrk)

        debugs = 4
        for row in tables[table_name]:
            erow = [_fmt(row.get(h)) for h in hdrk]
            try:
                wsh.append(erow)
            except (ValueError, openpyxl.utils.exceptions.IllegalCharacterError) as exc:
                debugs -= 1
                if debugs > 0:
                    _LOGGER.warning(
                        "Error writing %s, hdrs: %s - %s", list(erow), hdrk, exc
                    )
                wsh.append([str(row.get(h, "")) for h in hdrk])
        if debugs < 0:
            _LOGGER.warning("Total %s errors", 2 - debugs)

    wbk.save(_get_filename(file))  # Write to disk
